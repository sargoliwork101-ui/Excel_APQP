"""Control Plan merger using the CP template and station workbooks."""
from copy import copy
from pathlib import Path
import re
import openpyxl
from openpyxl.utils import get_column_letter

CP_SHEET = "برنامه کنترل  "
HEADER_ROWS = 9
DATA_START = 10
FOOTER_MARKERS = ("تهیه کنندگان", "تایید کننده", "تصویب کننده", "علامت", "بازنگری فرم", "کد فرم")


def _last_value_row(ws):
    for r in range(ws.max_row, 0, -1):
        if any(ws.cell(r, c).value is not None for c in range(1, min(ws.max_column, 60) + 1)):
            return r
    return ws.max_row


def _footer_row(ws):
    for r in range(DATA_START, _last_value_row(ws) + 1):
        values = " ".join(str(ws.cell(r, c).value or "") for c in range(1, min(ws.max_column, 20) + 1))
        if any(marker in values for marker in FOOTER_MARKERS):
            return r
    return _last_value_row(ws) + 1


def _copy_cell(src, dst):
    dst.value = src.value
    if src.has_style:
        dst.font = copy(src.font); dst.fill = copy(src.fill)
        dst.border = copy(src.border); dst.alignment = copy(src.alignment)
        dst.number_format = src.number_format; dst.protection = copy(src.protection)


def _copy_rows(src, dst, start, end, target, max_col):
    offset = target - start
    for r in range(start, end + 1):
        tr = r + offset
        if src.row_dimensions[r].height is not None:
            dst.row_dimensions[tr].height = src.row_dimensions[r].height
        for c in range(1, max_col + 1):
            _copy_cell(src.cell(r, c), dst.cell(tr, c))
    for merged in list(src.merged_cells.ranges):
        if merged.max_row < start or merged.min_row > end or merged.max_col > max_col:
            continue
        a = max(merged.min_row, start) + offset
        b = min(merged.max_row, end) + offset
        try:
            dst.merge_cells(f"{get_column_letter(merged.min_col)}{a}:{get_column_letter(merged.max_col)}{b}")
        except Exception:
            pass
    return end - start + 1


def _copy_sheet_properties(src, dst):
    dst.sheet_view.rightToLeft = src.sheet_view.rightToLeft
    dst.sheet_view.showGridLines = src.sheet_view.showGridLines
    dst.freeze_panes = src.freeze_panes
    dst.page_setup = copy(src.page_setup); dst.page_margins = copy(src.page_margins)
    for c in range(1, src.max_column + 1):
        letter = get_column_letter(c)
        if src.column_dimensions[letter].width is not None:
            dst.column_dimensions[letter].width = src.column_dimensions[letter].width


def merge_cp(template_path, station_paths, output_path, sheet_name=CP_SHEET):
    """Merge selected CP workbooks in ``station_paths`` in their given order."""
    twb = openpyxl.load_workbook(template_path, keep_links=False)
    actual = next((n for n in twb.sheetnames if n.strip().lower() == sheet_name.strip().lower()), None)
    if actual is None:
        raise ValueError(f"CP template sheet not found: {sheet_name}")
    tws = twb[actual]; max_col = tws.max_column
    out = openpyxl.Workbook(); out.remove(out.active)
    ws = out.create_sheet(actual); _copy_sheet_properties(tws, ws)
    _copy_rows(tws, ws, 1, HEADER_ROWS, 1, max_col)
    row = HEADER_ROWS + 1
    opened = {}
    for path in station_paths:
        path = str(path)
        if path not in opened: opened[path] = openpyxl.load_workbook(path, keep_links=False)
        wb = opened[path]
        source_name = next((n for n in wb.sheetnames if n.strip().lower() == actual.strip().lower()), wb.sheetnames[0])
        sws = wb[source_name]
        footer = _footer_row(sws)
        _copy_rows(sws, ws, DATA_START, footer - 1, row, max_col)
        row += footer - DATA_START
    footer_start = _footer_row(tws)
    if footer_start <= _last_value_row(tws):
        _copy_rows(tws, ws, footer_start, _last_value_row(tws), row, max_col)
    # Changes/history sheet is copied from the template, never fabricated.
    for name in twb.sheetnames:
        if name == actual: continue
        src = twb[name]; dst = out.create_sheet(name); _copy_sheet_properties(src, dst)
        _copy_rows(src, dst, 1, _last_value_row(src), 1, src.max_column)
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    out.save(output_path)
    for wb in opened.values(): wb.close()
    twb.close(); out.close()
    return str(output_path)
