"""
Read PFMEA-style workbooks and extract station structure.

A PFMEA file layout (as understood by this tool):

    Rows 1 .. header_rows              -> Header (company/product/column titles)
    Rows header_rows+1 .. footer_start -> Data rows (one or more stations)
    Rows footer_start .. end           -> Footer (signatures, distribution)

For every "station block" the first row has an OPC code in column A and
a station name in column B (A/B are often merged across the whole block).
"""
from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import List, Optional, Tuple
import openpyxl

from .config import MergeSettings


@dataclass
class StationBlock:
    """One station inside a source workbook (row range in that workbook)."""
    opc_code: str
    name: str
    start_row: int
    end_row: int
    source_file: str = ""
    order_index: int = 0
    failure_modes: List[str] = field(default_factory=list)

    @property
    def row_count(self) -> int:
        return self.end_row - self.start_row + 1

    @property
    def display_label(self) -> str:
        code = str(self.opc_code) if self.opc_code is not None else ""
        return f"{code} - {self.name}".strip(" -")

    @property
    def failure_mode_text(self) -> str:
        return "\n".join(self.failure_modes)


@dataclass
class WorkbookAnalysis:
    path: str
    sheet_name: str
    stations: List[StationBlock] = field(default_factory=list)
    footer_start_row: Optional[int] = None
    product_name: str = ""
    product_code: str = ""
    is_valid: bool = True
    error: str = ""


# ---------------------------------------------------------------------------

def _cell_str(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def last_non_empty_row(ws) -> int:
    """
    Return the row number of the last cell that actually contains a value.

    openpyxl's ws.max_row can be very optimistic — some files report 834
    when only the first 16 rows have real content. This walks upward until
    it finds a real value.
    """
    max_col = min(ws.max_column, 65)
    for r in range(ws.max_row, 0, -1):
        for c in range(1, max_col + 1):
            if ws.cell(row=r, column=c).value is not None:
                return r
    return ws.max_row


def _find_footer_start(ws, settings: MergeSettings, effective_max_row: int) -> Optional[int]:
    """
    Look for the first row in [data_start_row..effective_max_row] whose columns
    1..scan_cols contain any footer marker keyword.
    """
    markers = [m.strip() for m in settings.footer_markers if m and m.strip()]
    if not markers:
        return None
    scan_cols = min(ws.max_column, 20)
    for row in range(settings.data_start_row, effective_max_row + 1):
        for col in range(1, scan_cols + 1):
            val = _cell_str(ws.cell(row=row, column=col).value)
            if not val:
                continue
            for marker in markers:
                if marker in val:
                    return row
    return None


def _looks_like_opc(opc: str, max_len: int) -> bool:
    if not opc:
        return False
    if len(opc) > max_len:
        return False
    if any(ch in opc for ch in ("\n", "\r", ":")):
        return False
    # An OPC code is a short identifier, typically alphanumeric.
    # Allow at most one internal space (very rare); reject long text.
    if opc.count(" ") > 1:
        return False
    return True


def _find_stations(ws, settings: MergeSettings, data_end_row: int) -> List[StationBlock]:
    """
    Walk the OPC column between data_start_row and data_end_row and build
    StationBlock objects. A new station begins whenever a non-empty OPC
    code appears in column A; the block ends just before the next code.
    """
    opc_col = settings.opc_column
    name_col = settings.name_column
    max_len = getattr(settings, "max_opc_length", 20)

    starts: List[Tuple[int, str, str]] = []
    for row in range(settings.data_start_row, data_end_row + 1):
        opc = _cell_str(ws.cell(row=row, column=opc_col).value)
        if not _looks_like_opc(opc, max_len):
            continue
        name = _cell_str(ws.cell(row=row, column=name_col).value)
        starts.append((row, opc, name))

    stations: List[StationBlock] = []
    for i, (start_row, opc, name) in enumerate(starts):
        end_row = (starts[i + 1][0] - 1) if i + 1 < len(starts) else data_end_row
        # The UI must show the failure-mode cell belonging to the station
        # itself, not collect text from other rows or other sections. In the
        # standard PFMEA template this is column C on the station's first row.
        failure_value = _cell_str(ws.cell(row=start_row, column=3).value)
        failure_modes = [failure_value] if failure_value else []
        stations.append(StationBlock(
            opc_code=opc,
            name=name,
            start_row=start_row,
            end_row=end_row,
            order_index=i,
            failure_modes=failure_modes,
        ))
    return stations


def _extract_product_info(ws) -> Tuple[str, str]:
    """Best-effort extraction of product name/code from the header rows."""
    product_name = ""
    product_code = ""
    for row in range(1, 10):
        for col in range(1, min(ws.max_column, 30) + 1):
            val = _cell_str(ws.cell(row=row, column=col).value)
            if not val:
                continue
            if "نام قطعه" in val:
                after = val.split(":", 1)[-1].strip()
                if after:
                    product_name = after
            elif "شماره فنی" in val:
                after = val.split(":", 1)[-1].strip()
                if after:
                    product_code = after
    return product_name, product_code


def analyze_workbook(path: str | Path, settings: MergeSettings) -> WorkbookAnalysis:
    """Open a workbook and figure out its stations, footer, product info."""
    path = str(path)
    try:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=False)
    except Exception as e:
        return WorkbookAnalysis(
            path=path, sheet_name="", is_valid=False,
            error=f"cannot open: {e}",
        )

    # find sheet (case-insensitive fallback)
    sheet_name = settings.sheet_name
    if sheet_name not in wb.sheetnames:
        lower = sheet_name.strip().lower()
        for name in wb.sheetnames:
            if name.strip().lower() == lower:
                sheet_name = name
                break
        else:
            wb.close()
            return WorkbookAnalysis(
                path=path, sheet_name="", is_valid=False,
                error=f"sheet '{settings.sheet_name}' not found",
            )
    ws = wb[sheet_name]

    effective_max = last_non_empty_row(ws)
    footer_start = _find_footer_start(ws, settings, effective_max)
    data_end = (footer_start - 1) if footer_start else effective_max
    stations = _find_stations(ws, settings, data_end)
    for s in stations:
        s.source_file = path
    product_name, product_code = _extract_product_info(ws)

    analysis = WorkbookAnalysis(
        path=path,
        sheet_name=sheet_name,
        stations=stations,
        footer_start_row=footer_start,
        product_name=product_name,
        product_code=product_code,
        is_valid=len(stations) > 0,
        error="" if stations else "no stations detected",
    )
    wb.close()
    return analysis
