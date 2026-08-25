"""
Merge station blocks from several PFMEA workbooks into a single output
workbook that preserves the template's formatting (styles, merged cells,
column widths, row heights).

Output layout:

    [ template header rows           ]  (once, copied from template)
    [ station 1 data rows            ]  (copied from source file 1)
    [ station 2 data rows            ]  (copied from source file 2)
    ...
    [ template footer rows           ]  (once, copied from template)

The History sheet is copied verbatim from the template (never fabricated).
"""
from __future__ import annotations

from copy import copy
from pathlib import Path
from typing import Iterable, List, Optional, Tuple
import re
import tempfile
import zipfile
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .config import MergeSettings
from .excel_reader import StationBlock, WorkbookAnalysis, analyze_workbook, last_non_empty_row


# ---------------------------------------------------------------------------
# helpers to copy formatting

def _copy_cell(src_cell, dst_cell, row_offset: int = 0) -> None:
    """Copy a cell while remapping styles and relative formula references."""
    value = src_cell.value
    # Source station blocks are moved to a new row. Relative formulas (RPN,
    # SO, etc.) must follow them; copying the raw text would make every block
    # refer back to the source row numbers and produce incorrect PFMEA data.
    try:
        from openpyxl.formula.translate import Translator
        if isinstance(value, str) and value.startswith("=") and row_offset:
            value = Translator(value, origin=src_cell.coordinate).translate_formula(
                dst_cell.coordinate)
        elif hasattr(value, "text") and row_offset:
            # Preserve Excel array formulas while shifting their references.
            from openpyxl.worksheet.formula import ArrayFormula
            text = Translator(value.text, origin=src_cell.coordinate).translate_formula(
                dst_cell.coordinate)
            ref = getattr(value, "ref", dst_cell.coordinate)
            ref = re.sub(r"(\$?[A-Z]{1,3}\$?)(\d+)",
                         lambda m: f"{m.group(1)}{int(m.group(2)) + row_offset}", ref)
            value = ArrayFormula(ref=ref, text=text)
    except Exception:
        pass
    dst_cell.value = value
    if src_cell.has_style:
        dst_cell.font = copy(src_cell.font)
        dst_cell.fill = copy(src_cell.fill)
        dst_cell.border = copy(src_cell.border)
        dst_cell.alignment = copy(src_cell.alignment)
        dst_cell.number_format = src_cell.number_format
        dst_cell.protection = copy(src_cell.protection)
    if src_cell.hyperlink:
        try:
            dst_cell.hyperlink = copy(src_cell.hyperlink)
        except Exception:
            pass
    # Comments on merged cells break the file, so skip them.


def _copy_column_dims(src_ws: Worksheet, dst_ws: Worksheet, max_col: int) -> None:
    """Copy column widths / hidden / outline once (idempotent)."""
    for col_idx in range(1, max_col + 1):
        letter = get_column_letter(col_idx)
        src_dim = src_ws.column_dimensions.get(letter)
        if src_dim is None:
            continue
        dst_dim = dst_ws.column_dimensions[letter]
        if src_dim.width is not None:
            dst_dim.width = src_dim.width
        dst_dim.hidden = src_dim.hidden
        if src_dim.outlineLevel:
            dst_dim.outlineLevel = src_dim.outlineLevel
        if src_dim.bestFit:
            dst_dim.bestFit = src_dim.bestFit


def _copy_row_range(
    src_ws: Worksheet,
    dst_ws: Worksheet,
    src_start: int,
    src_end: int,
    dst_start: int,
    max_col: int,
) -> int:
    """
    Copy rows [src_start..src_end] from src_ws to dst_ws starting at
    dst_start. Copies values, styles, merged cells (relative), and row
    heights. Returns the number of rows copied.

    Merged cells that overlap the range are clipped to it. Any part of a
    merged range that falls outside the copied rows is dropped (this is
    what you want: it means the destination merge only covers the rows
    we actually kept).
    """
    row_count = src_end - src_start + 1
    offset = dst_start - src_start

    # cells + row heights
    for r in range(src_start, src_end + 1):
        dst_row = r + offset
        src_row_dim = src_ws.row_dimensions.get(r)
        if src_row_dim is not None:
            dst_row_dim = dst_ws.row_dimensions[dst_row]
            if src_row_dim.height is not None:
                dst_row_dim.height = src_row_dim.height
            if src_row_dim.hidden:
                dst_row_dim.hidden = True
            if src_row_dim.outlineLevel:
                dst_row_dim.outlineLevel = src_row_dim.outlineLevel
        for c in range(1, max_col + 1):
            _copy_cell(src_ws.cell(row=r, column=c),
                       dst_ws.cell(row=dst_row, column=c), row_offset=offset)

    # merged cells: keep any range whose bounds overlap [src_start..src_end]
    # and clip them into the destination range.
    for mr in list(src_ws.merged_cells.ranges):
        if mr.max_row < src_start or mr.min_row > src_end:
            continue
        if mr.min_col > max_col:
            continue
        clip_min_row = max(mr.min_row, src_start) + offset
        clip_max_row = min(mr.max_row, src_end) + offset
        clip_min_col = mr.min_col
        clip_max_col = min(mr.max_col, max_col)
        if clip_min_row > clip_max_row or clip_min_col > clip_max_col:
            continue
        if clip_min_row == clip_max_row and clip_min_col == clip_max_col:
            # single-cell "merge" is not useful; skip
            continue
        new_range = (
            f"{get_column_letter(clip_min_col)}{clip_min_row}:"
            f"{get_column_letter(clip_max_col)}{clip_max_row}"
        )
        try:
            dst_ws.merge_cells(new_range)
        except Exception:
            pass

    return row_count


def _copy_images(src_ws: Worksheet, dst_ws: Worksheet) -> None:
    """Copy header logos/images and their anchors from the template."""
    for image in getattr(src_ws, "_images", []):
        try:
            cloned = copy(image)
            cloned.anchor = copy(image.anchor)
            dst_ws.add_image(cloned)
        except Exception:
            # A non-critical drawing must never prevent the PFMEA merge.
            continue


def _copy_sheet_settings(src_ws: Worksheet, dst_ws: Worksheet) -> None:
    """Copy sheet-level formatting (page setup, freeze pane, RTL, print area)."""
    try:
        dst_ws.sheet_view.rightToLeft = src_ws.sheet_view.rightToLeft
    except Exception:
        pass
    try:
        dst_ws.sheet_view.showGridLines = src_ws.sheet_view.showGridLines
    except Exception:
        pass
    try:
        dst_ws.freeze_panes = src_ws.freeze_panes
    except Exception:
        pass
    try:
        dst_ws.page_setup = copy(src_ws.page_setup)
        dst_ws.page_margins = copy(src_ws.page_margins)
        dst_ws.print_options = copy(src_ws.print_options)
        dst_ws.sheet_properties = copy(src_ws.sheet_properties)
        dst_ws.print_title_rows = src_ws.print_title_rows
        dst_ws.print_title_cols = src_ws.print_title_cols
    except Exception:
        pass


def _copy_whole_sheet(src_ws: Worksheet, dst_ws: Worksheet) -> None:
    """Copy an entire worksheet 1:1 into dst_ws (which must be empty)."""
    max_col = src_ws.max_column
    last_row = _last_non_empty_row(src_ws) or src_ws.max_row
    _copy_sheet_settings(src_ws, dst_ws)
    _copy_column_dims(src_ws, dst_ws, max_col)
    _copy_row_range(src_ws, dst_ws, 1, last_row, 1, max_col)


# kept for backwards compat within this module
_last_non_empty_row = last_non_empty_row


def _rewrite_so_rpn_formulas(ws: Worksheet, data_start: int, data_end: int) -> None:
    """Write SO and RPN formulas explicitly for every merged data row."""
    from openpyxl.cell.cell import MergedCell
    for row in range(data_start, data_end + 1):
        # E = Severity, H = Occurrence, L = Detection, I = SO, M = RPN.
        # Do not trust formulas copied from source files: after rows are moved
        # their references can point to the old source row.
        so_cell = ws.cell(row=row, column=9)
        rpn_cell = ws.cell(row=row, column=13)
        # Some PFMEA templates merge the first station's descriptive row
        # across C:V, which makes I/M read-only MergedCell placeholders.
        if not isinstance(so_cell, MergedCell):
            so_cell.value = f"=H{row}*E{row}"
        if not isinstance(rpn_cell, MergedCell):
            rpn_cell.value = f"=L{row}*H{row}*E{row}"


def _update_rpn_formula(ws: Worksheet, data_start: int, data_end: int,
                        percent: int) -> None:
    """Update the template's AQ2 top-RPN formula for the merged data range."""
    if data_end < data_start:
        return
    cell = ws["AQ2"]
    old = cell.value
    # openpyxl 3.1 reads the template's AQ2 as an ArrayFormula object.
    # Work with its text while retaining the array-formula representation.
    old_text = getattr(old, "text", old)
    if isinstance(old_text, str) and old_text.startswith("="):
        formula = old_text
        # Keep the template's function/version, but never leave its stale
        # M10:M901 range in the merged workbook.
        # Preserve the template's first RPN row (the supplied template uses
        # row 10, while the station parser starts at row 9).
        match = re.search(r"\$M\$(\d+):\$M\$?\d+", formula)
        formula_start = int(match.group(1)) if match else data_start
        rng = f"$M${formula_start}:$M${data_end}"
        formula = re.sub(r"\$M\$\d+:\$M\$?\d+", rng, formula)
        # The template currently contains *0.2. Replace that multiplier while
        # retaining any future formula structure around it.
        formula = re.sub(r"\*\s*0(?:\.\d+)?", f"*{percent / 100:g}", formula,
                         count=1)
        if hasattr(old, "text"):
            try:
                from openpyxl.worksheet.formula import ArrayFormula
                cell.value = ArrayFormula(ref=getattr(old, "ref", "AQ2"), text=formula)
            except Exception:
                cell.value = formula
        else:
            cell.value = formula
    else:
        end = max(data_end, data_start)
        factor = percent / 100
        cell.value = (
            f"=INDEX(_xlfn._xlws.SORT($M${data_start}:$M${end},1,-1),"
            f"COUNT(_xlfn._xlws.SORT($M${data_start}:$M${end},1,-1))*{factor:g})"
        )


def _copy_conditional_formatting(src_ws: Worksheet, dst_ws: Worksheet,
                                 data_start: int, data_end: int) -> None:
    """Copy template conditional-format rules, expanding data rules to output."""
    if not src_ws.conditional_formatting:
        return
    from openpyxl.utils.cell import range_boundaries

    template_footer = None
    for cf in src_ws.conditional_formatting:
        for ref in str(cf.sqref).split():
            try:
                min_col, min_row, max_col, max_row = range_boundaries(ref)
            except ValueError:
                continue
            # Rules covering template data should cover all merged rows. Rules
            # outside the data area (e.g. header rules) retain their location.
            if min_row >= data_start and max_row >= data_start:
                # Header rows stay fixed; preserve the template's first data
                # row (M10 in the supplied template) and extend the rule to
                # the final merged data row.
                max_row = data_end
            new_ref = (f"{get_column_letter(min_col)}{min_row}:"
                       f"{get_column_letter(max_col)}{max_row}")
            for rule in cf.rules:
                copied_rule = copy(rule)
                if getattr(rule, "dxf", None) is not None:
                    copied_rule.dxf = copy(rule.dxf)
                dst_ws.conditional_formatting.add(new_ref, copied_rule)


def _attach_template_drawing(output_path: str, template_path: str) -> None:
    """Attach unsupported template drawings (logo/shapes) to the output ZIP.

    openpyxl does not load the legacy/grouped drawing used by this PFMEA
    header, so copying cells alone silently drops the logo. The drawing XML
    and its image relationship are safe to carry over unchanged because the
    output sheet keeps the same header coordinates.
    """
    try:
        with zipfile.ZipFile(template_path, "r") as src, zipfile.ZipFile(output_path, "r") as out:
            names = set(src.namelist())
            drawing_names = [n for n in names if n.startswith("xl/drawings/drawing") and n.endswith(".xml")]
            if not drawing_names:
                return
            drawing = sorted(drawing_names)[0]
            drawing_rels = drawing.replace("xl/drawings/", "xl/drawings/_rels/") + ".rels"
            media_names = [n for n in names if n.startswith("xl/media/")]
            sheet_name = "xl/worksheets/sheet1.xml"
            sheet_rels = "xl/worksheets/_rels/sheet1.xml.rels"
            sheet_xml = out.read(sheet_name).decode("utf-8")
            if "xmlns:r=" not in sheet_xml:
                sheet_xml = sheet_xml.replace(
                    "<worksheet ",
                    '<worksheet xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships" ',
                    1,
                )
            if "<drawing " not in sheet_xml:
                sheet_xml = sheet_xml.replace("</worksheet>", '<drawing r:id="rId2"/></worksheet>')
            rels_xml = ('<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
                        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
                        '<Relationship Id="rId2" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/drawing" '
                        'Target="../drawings/drawing1.xml"/></Relationships>')
            content = out.read("[Content_Types].xml").decode("utf-8")
            if 'Extension="png"' not in content:
                content = content.replace("</Types>", '<Default Extension="png" ContentType="image/png"/></Types>')
            if 'PartName="/xl/drawings/drawing1.xml"' not in content:
                content = content.replace("</Types>", '<Override PartName="/xl/drawings/drawing1.xml" ContentType="application/vnd.openxmlformats-officedocument.drawing+xml"/></Types>')
            with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx", dir=str(Path(output_path).parent)) as tmp:
                temp_name = tmp.name
            try:
                with zipfile.ZipFile(temp_name, "w", zipfile.ZIP_DEFLATED) as new:
                    has_sheet_rels = False
                    for item in out.infolist():
                        data = sheet_xml.encode() if item.filename == sheet_name else content.encode() if item.filename == "[Content_Types].xml" else out.read(item.filename)
                        if item.filename == sheet_rels:
                            has_sheet_rels = True
                            data = rels_xml.encode()
                        new.writestr(item, data)
                    if not has_sheet_rels:
                        new.writestr(sheet_rels, rels_xml.encode())
                    new.writestr(drawing, src.read(drawing))
                    if drawing_rels in names:
                        new.writestr(drawing_rels, src.read(drawing_rels))
                    for media in media_names:
                        new.writestr(media, src.read(media))
                Path(temp_name).replace(output_path)
            finally:
                Path(temp_name).unlink(missing_ok=True)
    except Exception:
        # A logo is valuable, but must not make a valid PFMEA output fail.
        return


# ---------------------------------------------------------------------------
# public API

def merge_pfmea(
    template_path: str | Path,
    station_selections: List[Tuple[str, StationBlock]],
    output_path: str | Path,
    settings: MergeSettings,
    merge_history: bool = True,
    progress_cb=None,
) -> str:
    """
    Build the output workbook from a template + selected station blocks.
    station_selections is an ordered list of (source_file_path, StationBlock).
    """
    template_path = str(template_path)
    output_path = str(output_path)

    def report(pct: int, msg: str = ""):
        if progress_cb:
            try:
                progress_cb(pct, msg)
            except Exception:
                pass

    report(1, "Loading template...")
    template_wb = openpyxl.load_workbook(template_path, keep_links=False)

    tpl_sheet_name = settings.sheet_name
    if tpl_sheet_name not in template_wb.sheetnames:
        raise ValueError(f"Template does not contain sheet '{tpl_sheet_name}'")
    tpl_ws = template_wb[tpl_sheet_name]

    tpl_analysis = analyze_workbook(template_path, settings)
    tpl_footer_start = tpl_analysis.footer_start_row
    tpl_footer_end = _last_non_empty_row(tpl_ws)
    max_col = tpl_ws.max_column

    # ---- build the output workbook and copy header ------------------------
    out_wb = openpyxl.Workbook()
    out_wb.remove(out_wb.active)
    # Let Excel recalculate AQ2 (and other formulas) as soon as the output is
    # opened instead of leaving a stale cached value from the template.
    try:
        out_wb.calculation.fullCalcOnLoad = True
        out_wb.calculation.forceFullCalc = True
        out_wb.calculation.calcMode = "auto"
    except Exception:
        pass
    out_ws = out_wb.create_sheet(title=tpl_sheet_name)

    _copy_sheet_settings(tpl_ws, out_ws)
    _copy_column_dims(tpl_ws, out_ws, max_col)
    _copy_images(tpl_ws, out_ws)

    report(5, "Copying header...")
    _copy_row_range(tpl_ws, out_ws, 1, settings.header_rows, 1, max_col)

    # ---- copy each selected station ---------------------------------------
    write_row = settings.header_rows + 1
    n = max(1, len(station_selections))
    src_cache: dict[str, openpyxl.Workbook] = {}
    for i, (src_path, block) in enumerate(station_selections, start=1):
        report(5 + int(80 * i / n),
               f"Merging station {i}/{n}: {block.display_label}")
        if src_path not in src_cache:
            src_cache[src_path] = openpyxl.load_workbook(src_path, keep_links=False)
        src_wb = src_cache[src_path]
        sheet = tpl_sheet_name if tpl_sheet_name in src_wb.sheetnames \
            else src_wb.sheetnames[0]
        src_ws = src_wb[sheet]
        rows = _copy_row_range(
            src_ws, out_ws,
            block.start_row, block.end_row,
            write_row, max_col,
        )
        write_row += rows

    for wb in src_cache.values():
        wb.close()

    # ---- copy footer ------------------------------------------------------
    if tpl_footer_start:
        report(88, "Copying footer...")
        _copy_row_range(
            tpl_ws, out_ws,
            tpl_footer_start, tpl_footer_end,
            write_row, max_col,
        )

    # AQ2 is a template formula, so its source range and percentage must be
    # recalculated for the rows that actually made it into this output.
    data_end = write_row - 1
    _rewrite_so_rpn_formulas(out_ws, settings.data_start_row, data_end)
    _update_rpn_formula(
        out_ws, settings.data_start_row, data_end,
        max(1, min(100, int(getattr(settings, "rpn_top_percent", 20)))),
    )
    # Conditional-format rules refer to differential styles by dxfId. Since
    # the output workbook is created from scratch, copy that style table too;
    # otherwise the rules can exist in XML but render with no formatting.
    try:
        for dxf in template_wb._differential_styles.styles:
            out_wb._differential_styles.add(copy(dxf))
    except Exception:
        pass
    _copy_conditional_formatting(
        tpl_ws, out_ws, settings.data_start_row, data_end,
    )

    # ---- History sheet ---------------------------------------------------
    # We copy it VERBATIM from the template (never fabricate or "merge" it
    # from every source file — that produced noisy duplicates before).
    if merge_history:
        report(92, "Copying history sheet from template...")
        _copy_template_history(template_wb, out_wb, settings)

    report(97, "Saving output...")
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    try:
        out_wb.save(output_path)
    except PermissionError as e:
        template_wb.close()
        out_wb.close()
        raise PermissionError(
            f"Could not write to '{output_path}'. "
            f"Is the file open in Excel? Please close it and try again.\n"
            f"({e})"
        ) from e
    template_wb.close()
    out_wb.close()
    _attach_template_drawing(output_path, template_path)
    report(100, "Done.")
    return output_path


def _copy_template_history(template_wb, out_wb, settings: MergeSettings) -> None:
    """Copy the template's History sheet verbatim, if one exists."""
    tpl_hist_name = None
    for name in template_wb.sheetnames:
        if name.strip().lower() == settings.history_sheet.strip().lower():
            tpl_hist_name = name
            break
    if tpl_hist_name is None:
        return
    src_ws = template_wb[tpl_hist_name]
    dst_ws = out_wb.create_sheet(title=tpl_hist_name)
    _copy_whole_sheet(src_ws, dst_ws)
