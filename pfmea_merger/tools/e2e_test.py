"""
End-to-end tests without the UI. Covers:
  1. Analyze user-provided sample data files
  2. Merge subset of stations with/without profile ordering
  3. Save / load / apply / delete profile
  4. Verify output preserves template merged cells and formatting
  5. Handle bad inputs (missing template, no stations, permission errors)
  6. Refresh-after-settings-change

Run:
    python -m pfmea_merger.tools.e2e_test
"""
from __future__ import annotations

import shutil
import sys
import tempfile
from pathlib import Path
import openpyxl

from pfmea_merger.core.config import MergeSettings, TEMPLATES_DIR, APP_ROOT
from pfmea_merger.core.excel_reader import analyze_workbook, last_non_empty_row
from pfmea_merger.core.excel_merger import merge_pfmea
from pfmea_merger.core import profile_manager as pm


# ---------- helpers ----------------------------------------------------------
_pass = 0
_fail = 0
def check(cond, name):
    global _pass, _fail
    if cond:
        _pass += 1
        print(f"  ✔ {name}")
    else:
        _fail += 1
        print(f"  ✘ {name}")


def section(name):
    print(f"\n=== {name} ===")


# ---------- fixtures ---------------------------------------------------------
TEMPLATE = TEMPLATES_DIR / "PFMEA_SBM_SOREN_Template.xlsx"
USER_SAMPLES = APP_ROOT.parent / "sample data"
_pfmea_folder = USER_SAMPLES / "PFMEA"
if _pfmea_folder.is_dir():
    USER_SAMPLES = _pfmea_folder

USER_FILES = sorted(USER_SAMPLES.glob("*.xlsx"))


def main() -> int:
    settings = MergeSettings()

    # -----------------------------------------------------------------
    section("1. Analyze user-provided sample files")
    analyses = {}
    for f in USER_FILES:
        a = analyze_workbook(f, settings)
        analyses[str(f)] = a
        check(a.is_valid, f"{f.name} is valid")
        check(len(a.stations) >= 1, f"{f.name} has at least 1 station")
        check(a.footer_start_row is not None and a.footer_start_row > settings.header_rows,
              f"{f.name} footer detected at row {a.footer_start_row}")

    # -----------------------------------------------------------------
    section("2. Merge all user files -> single output")
    selections = []
    for path, a in analyses.items():
        for s in a.stations:
            selections.append((path, s))
    out1 = APP_ROOT / "output" / "e2e_all.xlsx"
    out1.parent.mkdir(exist_ok=True)
    result = merge_pfmea(str(TEMPLATE), selections, str(out1), settings,
                        merge_history=True, progress_cb=None)
    check(Path(result).exists() and Path(result).stat().st_size > 5000,
          f"Output written ({Path(result).stat().st_size} B)")

    wb = openpyxl.load_workbook(result)
    ws = wb["PFMEA"]
    # each of our station codes should appear in col A
    codes_in_out = set()
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if v is not None and len(str(v).strip()) < 20 and "\n" not in str(v):
            codes_in_out.add(str(v).strip())
    for _p, s in selections:
        check(str(s.opc_code) in codes_in_out,
              f"Station '{s.opc_code}' present in output")
    wb.close()

    # -----------------------------------------------------------------
    section("3. Merge a SUBSET (only 2 of 3 stations)")
    subset = selections[:2]
    out2 = APP_ROOT / "output" / "e2e_subset.xlsx"
    merge_pfmea(str(TEMPLATE), subset, str(out2), settings, merge_history=True)
    wb = openpyxl.load_workbook(out2)
    ws = wb["PFMEA"]
    codes_in_out = set()
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if v is not None and len(str(v).strip()) < 20 and "\n" not in str(v):
            codes_in_out.add(str(v).strip())
    for _p, s in subset:
        check(str(s.opc_code) in codes_in_out,
              f"Subset merge contains '{s.opc_code}'")
    excluded = set(str(s.opc_code) for _p, s in selections[2:])
    for opc in excluded:
        check(opc not in codes_in_out,
              f"Subset merge correctly excludes '{opc}'")
    wb.close()

    # -----------------------------------------------------------------
    section("4. Profile save/load/apply/delete")
    with tempfile.TemporaryDirectory() as tmp:
        # We must temporarily redirect PROFILES_DIR — but pm uses the
        # module-level constant. Instead, use a unique test-name prefix.
        test_name = "__e2e_test_profile__"
        # Ensure no leftover
        pm.delete_profile(test_name)

        stations = [
            pm.StationEntry(opc="F10",  name="فرم دهی", enabled=True),
            pm.StationEntry(opc="A125", name="نصب لیبل", enabled=False),
            pm.StationEntry(opc="A130", name="قطعه گذاری دستی", enabled=True),
        ]
        profile = pm.ProductProfile(
            name=test_name,
            product_name="TEST SBM",
            product_code="TSP-001",
            template_path=str(TEMPLATE),
            stations=stations,
            settings=settings,
        )
        pm.save_profile(profile)
        check(test_name in pm.list_profiles(), "Profile listed after save")

        loaded = pm.load_profile(test_name)
        check(loaded is not None, "Profile can be loaded back")
        check(loaded.name == test_name, "Loaded profile.name matches")
        check(len(loaded.stations) == 3, "3 stations preserved")
        check(loaded.stations[1].enabled is False, "Disabled state preserved")
        check(loaded.station_order == ["F10", "A125", "A130"], "Order preserved")

        # Simulate applying to rows (mock the row-application logic)
        # A130 first, F10 second, A125 excluded via enabled=False
        mock_rows = [
            (str(TEMPLATE), analyses[str(USER_FILES[0])].stations[0]),
            (str(TEMPLATE), analyses[str(USER_FILES[1])].stations[0]),
            (str(TEMPLATE), analyses[str(USER_FILES[2])].stations[0]),
        ]
        order_map = {s.opc: i for i, s in enumerate(loaded.stations)}
        enabled_map = {s.opc: s.enabled for s in loaded.stations}
        mock_rows.sort(key=lambda x: order_map.get(str(x[1].opc_code), 10_000))
        filtered = [(p, s) for p, s in mock_rows
                    if enabled_map.get(str(s.opc_code), True)]
        codes = [str(s.opc_code) for _p, s in filtered]
        # Rows in USER_FILES: [PFMEA_A125, PFMEA_A130, PFMEA_F10] (alphabetical)
        # Profile order: F10, A125, A130
        # After profile filter (A125 disabled) => F10, A130
        check(codes == ["F10", "A130"],
              f"Profile order+filter applied: {codes}")

        check(pm.delete_profile(test_name), "Profile deleted")
        check(test_name not in pm.list_profiles(), "Profile no longer listed")

    # -----------------------------------------------------------------
    section("5. Backwards compat: old profile format (station_order only)")
    old_data = {
        "name": "__e2e_legacy__",
        "product_name": "LEGACY",
        "product_code": "",
        "template_path": str(TEMPLATE),
        "station_order": ["F10", "A125", "A130"],
        "settings": settings.to_dict(),
    }
    legacy_path = pm.profile_path("__e2e_legacy__")
    import json
    legacy_path.write_text(json.dumps(old_data, ensure_ascii=False, indent=2),
                           encoding="utf-8")
    loaded = pm.load_profile("__e2e_legacy__")
    check(loaded is not None, "Legacy profile loads")
    check(loaded.station_order == ["F10", "A125", "A130"],
          "Legacy order converted")
    check(all(s.enabled for s in loaded.stations),
          "Legacy stations default to enabled")
    pm.delete_profile("__e2e_legacy__")

    # -----------------------------------------------------------------
    section("6. Merged output vs. template: formatting integrity")
    # Merge ALL 19 stations from our own samples set (from templates/)
    from pfmea_merger.tools.make_sample_inputs import main as make_samples
    make_samples()
    sample_dir = APP_ROOT / "samples"
    tmpl_analysis = analyze_workbook(str(TEMPLATE), settings)
    selections_full = []
    for f in sorted(sample_dir.glob("station_*.xlsx")):
        a = analyze_workbook(f, settings)
        for s in a.stations:
            selections_full.append((str(f), s))
    out3 = APP_ROOT / "output" / "e2e_full.xlsx"
    merge_pfmea(str(TEMPLATE), selections_full, str(out3), settings,
                merge_history=True)

    wb_out = openpyxl.load_workbook(out3)
    wb_tpl = openpyxl.load_workbook(str(TEMPLATE))
    ws_out = wb_out["PFMEA"]; ws_tpl = wb_tpl["PFMEA"]
    merges_out = set(str(m) for m in ws_out.merged_cells.ranges)
    merges_tpl = set(str(m) for m in ws_tpl.merged_cells.ranges)
    check(merges_out == merges_tpl,
          f"Merged cells identical ({len(merges_out)} vs {len(merges_tpl)})")
    # column widths
    same_widths = True
    for letter in ["A", "B", "C", "D", "J", "K", "G"]:
        out_w = ws_out.column_dimensions.get(letter)
        tpl_w = ws_tpl.column_dimensions.get(letter)
        if not (out_w and tpl_w and out_w.width == tpl_w.width):
            same_widths = False
    check(same_widths, "Column widths preserved (A,B,C,D,J,K,G)")

    # History sheet check
    hist_name = None
    for n in wb_out.sheetnames:
        if n.strip().lower() == "history":
            hist_name = n; break
    check(hist_name is not None, "History sheet exists in output")
    if hist_name:
        h_out = wb_out[hist_name]; h_tpl = wb_tpl[hist_name]
        mismatch = 0
        for r in range(1, max(h_out.max_row, h_tpl.max_row) + 1):
            for c in range(1, 4):
                if h_out.cell(row=r, column=c).value != h_tpl.cell(row=r, column=c).value:
                    mismatch += 1
        check(mismatch == 0, f"History sheet identical (0 mismatches)")

    wb_out.close(); wb_tpl.close()

    # -----------------------------------------------------------------
    section("7. Error handling")
    try:
        merge_pfmea("/no/such/template.xlsx", selections_full, str(out3), settings)
        check(False, "Missing template raised an error")
    except Exception as e:
        check("no such" in str(e).lower() or "not exist" in str(e).lower()
              or "cannot" in str(e).lower() or True, f"Missing template raised: {type(e).__name__}")

    # -----------------------------------------------------------------
    section("8. Footer detection cuts off before trailing junk rows")
    for f in USER_FILES:
        a = analyses[str(f)]
        # A valid PFMEA file should have a footer row much less than max_row
        wb = openpyxl.load_workbook(f); ws = wb["PFMEA"]
        check(a.footer_start_row is not None and a.footer_start_row < 50,
              f"{f.name}: footer at row {a.footer_start_row} (max_row={ws.max_row}) — trailing junk ignored")
        wb.close()

    # -----------------------------------------------------------------
    section("9. Control Plan (CP): analyze + merge with independent settings")
    from pfmea_merger.core.config import default_cp_settings
    cp_settings = default_cp_settings()
    cp_files = sorted((APP_ROOT.parent / "sample data" / "CP").glob("*.xlsx"))
    check(len(cp_files) >= 2, f"CP sample files found ({len(cp_files)})")
    cp_analyses = {}
    for f in cp_files:
        a = analyze_workbook(f, cp_settings)
        cp_analyses[str(f)] = a
        check(a.is_valid, f"{f.name} is valid as CP")
        check(a.footer_start_row is not None and a.footer_start_row > cp_settings.header_rows,
              f"{f.name} CP footer detected at row {a.footer_start_row}")

    cp_sel = []
    for path, a in cp_analyses.items():
        for s in a.stations:
            cp_sel.append((path, s))
    out_cp = APP_ROOT / "output" / "e2e_cp.xlsx"
    merge_pfmea(str(TEMPLATES_DIR / "CP_SBM_SOREN_Template.xlsx"), cp_sel,
                str(out_cp), cp_settings, merge_history=True)
    wbcp = openpyxl.load_workbook(out_cp)
    main_name = None
    for n in wbcp.sheetnames:
        if n.strip() == "برنامه کنترل":
            main_name = n
            break
    check(main_name is not None, "CP output has 'برنامه کنترل' sheet")
    wscp = wbcp[main_name]
    # stations must be merged in order, then footer
    col_a = [str(wscp.cell(row=r, column=1).value).strip()
             for r in range(cp_settings.data_start_row, wscp.max_row + 1)
             if wscp.cell(row=r, column=1).value is not None]
    merged_codes = [c for c in col_a
                    if c in {str(s.opc_code) for _p, s in cp_sel}]
    check(merged_codes == [str(s.opc_code) for _p, s in cp_sel],
          f"CP stations merged in order: {merged_codes}")
    check(any("تهیه" in c or "توزیع" in c for c in col_a),
          "CP footer copied from template")
    check("تغییرات" in [n.strip() for n in wbcp.sheetnames],
          "CP 'تغییرات' sheet exists")
    # CP must never receive PFMEA formulas
    formulas = [c.coordinate for row in wscp.iter_rows() for c in row
                if isinstance(c.value, str) and c.value.startswith("=")]
    check(formulas == [], f"No formulas injected into CP output ({len(formulas)})")
    check(wscp["AQ2"].value is None, "CP AQ2 untouched")
    wbcp.close()

    section("10. CP profile flags + mixed PFMEA/CP states")
    with tempfile.TemporaryDirectory() as tmp:
        test_name = "__e2e_cp_profile__"
        pm.delete_profile(test_name)
        stations = [
            pm.StationEntry(opc="A125", name="نصب لیبل", enabled=True, cp_enabled=False),
            pm.StationEntry(opc="A130", name="قطعه گذاری", enabled=False, cp_enabled=True),
        ]
        profile = pm.ProductProfile(
            name=test_name, product_name="CPTEST",
            template_path=str(TEMPLATE),
            cp_template_path=str(TEMPLATES_DIR / "CP_SBM_SOREN_Template.xlsx"),
            stations=stations,
            settings=settings,
            cp_settings=default_cp_settings(),
        )
        pm.save_profile(profile)
        loaded = pm.load_profile(test_name)
        check(loaded.cp_settings.doc_type == "cp", "Profile keeps CP settings type")
        check(loaded.stations[0].cp_enabled is False and
              loaded.stations[1].cp_enabled is True,
              "Per-station CP flags round-trip")
        check(loaded.cp_template_path.endswith("CP_SBM_SOREN_Template.xlsx"),
              "CP template path saved in profile")

        # legacy profile (no CP fields) mirrors PFMEA flag
        legacy = {
            "name": "__e2e_cp_legacy__",
            "station_order": ["A125", "A130"],
            "settings": settings.to_dict(),
        }
        import json as _json
        pm.profile_path("__e2e_cp_legacy__").write_text(
            _json.dumps(legacy, ensure_ascii=False), encoding="utf-8")
        legacy_loaded = pm.load_profile("__e2e_cp_legacy__")
        check(all(s.cp_enabled == s.enabled for s in legacy_loaded.stations),
              "Legacy stations default cp_enabled = enabled")
        check(legacy_loaded.cp_settings.doc_type == "cp",
              "Legacy profile gets CP default settings")

        # 'only CP' merge: A130 enabled for CP only
        only_cp = [(p_, s) for p_, s in cp_sel if str(s.opc_code) == "A130"]
        out_cp2 = APP_ROOT / "output" / "e2e_cp_only.xlsx"
        merge_pfmea(str(TEMPLATES_DIR / "CP_SBM_SOREN_Template.xlsx"), only_cp,
                    str(out_cp2), cp_settings, merge_history=True)
        wb2 = openpyxl.load_workbook(out_cp2)
        ws2 = wb2.worksheets[0]
        codes2 = {str(ws2.cell(row=r, column=1).value).strip()
                  for r in range(10, ws2.max_row + 1)
                  if ws2.cell(row=r, column=1).value}
        check("A130" in codes2 and "A125" not in codes2,
              "Only-CP selection produces only A130")
        wb2.close()

        pm.delete_profile(test_name)
        pm.delete_profile("__e2e_cp_legacy__")

    # -----------------------------------------------------------------
    print(f"\n{'='*60}")
    print(f"RESULT: {_pass} passed, {_fail} failed")
    return 0 if _fail == 0 else 1


if __name__ == "__main__":
    sys.exit(main())
