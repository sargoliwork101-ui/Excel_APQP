"""
Split the SBM SOREN template into per-station sample input files.

Approach:
    For each station block we start from an EMPTY workbook, copy the
    template's header + footer + column dimensions + sheet settings, then
    copy just the rows belonging to that station. This preserves merges
    correctly (openpyxl.delete_rows does not shift merged ranges, so any
    approach based on deletion produces broken samples).

Usage:
    python -m pfmea_merger.tools.make_sample_inputs
"""
from __future__ import annotations

from pathlib import Path
import openpyxl

from pfmea_merger.core.config import MergeSettings, TEMPLATES_DIR, APP_ROOT
from pfmea_merger.core.excel_reader import analyze_workbook
from pfmea_merger.core.excel_merger import (
    _copy_row_range, _copy_column_dims, _copy_sheet_settings, _last_non_empty_row,
)


SAMPLES_DIR = APP_ROOT / "samples"
SAMPLES_DIR.mkdir(parents=True, exist_ok=True)

TEMPLATE = TEMPLATES_DIR / "PFMEA_SBM_SOREN_Template.xlsx"


def _safe_name(s: str) -> str:
    keep = "-_. "
    return "".join(c if (c.isalnum() or c in keep) else "_" for c in s).strip() or "station"


def main():
    settings = MergeSettings()
    analysis = analyze_workbook(TEMPLATE, settings)
    if not analysis.is_valid:
        print("Template invalid:", analysis.error)
        return
    stations = analysis.stations
    footer_start = analysis.footer_start_row

    tpl_wb = openpyxl.load_workbook(TEMPLATE)
    tpl_ws = tpl_wb[settings.sheet_name]
    max_col = tpl_ws.max_column
    tpl_footer_end = _last_non_empty_row(tpl_ws)

    # try to copy the History sheet as-is if it exists
    hist_name = None
    for name in tpl_wb.sheetnames:
        if name.strip().lower() == settings.history_sheet.strip().lower():
            hist_name = name
            break

    print(f"Found {len(stations)} stations, header={settings.header_rows}, "
          f"footer_start={footer_start}, footer_end={tpl_footer_end}, "
          f"max_col={max_col}")

    for i, block in enumerate(stations):
        out = SAMPLES_DIR / f"station_{i+1:02d}_{block.opc_code}_{_safe_name(block.name)[:20]}.xlsx"

        out_wb = openpyxl.Workbook()
        out_wb.remove(out_wb.active)
        out_ws = out_wb.create_sheet(title=settings.sheet_name)

        _copy_sheet_settings(tpl_ws, out_ws)
        _copy_column_dims(tpl_ws, out_ws, max_col)

        # 1) header
        _copy_row_range(tpl_ws, out_ws, 1, settings.header_rows, 1, max_col)
        write_row = settings.header_rows + 1

        # 2) just this station's rows
        rows = _copy_row_range(
            tpl_ws, out_ws,
            block.start_row, block.end_row,
            write_row, max_col,
        )
        write_row += rows

        # 3) footer
        if footer_start:
            _copy_row_range(tpl_ws, out_ws,
                            footer_start, tpl_footer_end,
                            write_row, max_col)

        # 4) copy History sheet as-is
        if hist_name:
            src_hist = tpl_wb[hist_name]
            dst_hist = out_wb.create_sheet(title=hist_name)
            _copy_sheet_settings(src_hist, dst_hist)
            hist_max_col = max(src_hist.max_column, 3)
            _copy_column_dims(src_hist, dst_hist, hist_max_col)
            _copy_row_range(src_hist, dst_hist, 1, src_hist.max_row, 1, hist_max_col)

        out_wb.save(out)
        out_wb.close()
        print(f"  ✔ {out.name}  ({block.opc_code} - {block.name}, {block.row_count} rows)")

    tpl_wb.close()


if __name__ == "__main__":
    main()
